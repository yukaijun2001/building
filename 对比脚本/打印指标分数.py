#!/usr/bin/env python3
"""
遍历 results 目录，导出每个实验指标到 xlsx。
并额外输出 pred_len_avg sheet：
- 24/36/48/60 各算一个平均（6个数据集：assembly/education/food/office/parking/public）
"""

from pathlib import Path
import argparse
import re
import numpy as np
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


METRIC_NAMES = ["MAE", "MSE", "RMSE", "MAPE", "MSPE"]
DATASETS = ["assembly", "education", "food", "office", "parking", "public"]
PRED_LENS = [24, 48, 72, 96]


def autosize_columns(ws, min_width=10, max_width=80):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def safe_float(x) -> Optional[float]:
    if isinstance(x, (int, float)) and np.isfinite(x):
        return float(x)
    return None


def safe_mean(values) -> Optional[float]:
    vals = [v for v in values if v is not None and np.isfinite(v)]
    return float(np.mean(vals)) if vals else None


def parse_dataset_and_predlen(exp_name: str):
    low = exp_name.lower()
    m_ds = re.search(r"_(assembly|education|food|office|parking|public)_", low)
    m_pl = re.search(r"_pl(\d+)", low)
    ds = m_ds.group(1) if m_ds else None
    pl = int(m_pl.group(1)) if m_pl else None
    return ds, pl


def compute_r2_from_dir(exp_dir: Path) -> Optional[float]:
    candidates = [("true.npy", "pred.npy"), ("trues.npy", "preds.npy")]
    for tname, pname in candidates:
        tp = exp_dir / tname
        pp = exp_dir / pname
        if tp.exists() and pp.exists():
            try:
                y_true = np.asarray(np.load(tp)).reshape(-1)
                y_pred = np.asarray(np.load(pp)).reshape(-1)
                if y_true.size == 0:
                    return None
                ss_res = np.sum((y_true - y_pred) ** 2)
                ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
                if ss_tot == 0:
                    return None
                return float(1.0 - ss_res / ss_tot)
            except Exception:
                return None
    return None


def write_metrics_sheet(wb: Workbook, rows: List[Dict[str, Any]]):
    ws = wb.active
    ws.title = "metrics"

    headers = ["实验目录", "MSE", "MAE", "MAPE", "R2", "RMSE", "MSPE", "metrics.npy 路径"]
    ws.append(headers)

    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r.get("exp_dir", ""),
            r.get("mse", None),
            r.get("mae", None),
            r.get("mape", None),
            r.get("r2", None),
            r.get("rmse", None),
            r.get("mspe", None),
            r.get("metrics_path", ""),
        ])

    # 可选：全体实验平均
    ws.append([
        "平均(全部实验)",
        safe_mean([r.get("mse") for r in rows]),
        safe_mean([r.get("mae") for r in rows]),
        safe_mean([r.get("mape") for r in rows]),
        safe_mean([r.get("r2") for r in rows]),
        None, None, ""
    ])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in [2, 3, 4, 5, 6, 7]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"

    autosize_columns(ws)


def write_pred_len_avg_sheet(wb: Workbook, rows: List[Dict[str, Any]]):
    if "pred_len_avg" in wb.sheetnames:
        wb.remove(wb["pred_len_avg"])
    ws = wb.create_sheet("pred_len_avg")

    headers = ["Pred_len", "Count(rows)", "Datasets_covered", "Avg_MSE", "Avg_MAE", "Avg_MAPE", "Avg_R2"]
    ws.append(headers)

    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for pl in PRED_LENS:
        pick = [r for r in rows if r.get("pred_len") == pl and r.get("dataset") in DATASETS]
        ds_cov = sorted({r.get("dataset") for r in pick if r.get("dataset")})
        ws.append([
            pl,
            len(pick),
            ",".join(ds_cov),
            safe_mean([r.get("mse") for r in pick]),
            safe_mean([r.get("mae") for r in pick]),
            safe_mean([r.get("mape") for r in pick]),
            safe_mean([r.get("r2") for r in pick]),
        ])

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in [4, 5, 6, 7]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"

    autosize_columns(ws)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--results_root",
        type=str,
        default="/home/ykj/build/Time-Series-Library/results",
        help="results 根目录",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="/home/ykj/build/对比脚本/表格/metrics_summary.xlsx",
        help="xlsx 输出路径",
    )
    args = parser.parse_args()

    results_root = Path(args.results_root)
    if not results_root.exists():
        print(f"结果目录不存在: {results_root}")
        return

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"{'实验目录':<100} | {'MSE':<10} | {'MAE':<10}")
    print("-" * 125)

    rows = []
    for exp_dir in sorted(results_root.iterdir()):
        if not exp_dir.is_dir():
            continue

        metrics_path = exp_dir / "metrics.npy"
        if not metrics_path.exists():
            continue

        try:
            metrics = np.load(metrics_path)

            values = {name.lower(): None for name in METRIC_NAMES}
            for i, name in enumerate(METRIC_NAMES):
                if i < len(metrics):
                    values[name.lower()] = float(metrics[i])

            ds, pl = parse_dataset_and_predlen(exp_dir.name)
            r2 = compute_r2_from_dir(exp_dir)

            mse = safe_float(values["mse"])
            mae = safe_float(values["mae"])

            if mse is not None and mae is not None:
                print(f"{exp_dir.name:<100} | {mse:<10.6f} | {mae:<10.6f}")
            else:
                print(f"{exp_dir.name:<100} | {'NA':<10} | {'NA':<10}")

            rows.append({
                "exp_dir": exp_dir.name,
                "dataset": ds,
                "pred_len": pl,
                "mse": mse,
                "mae": mae,
                "mape": safe_float(values["mape"]),
                "r2": safe_float(r2),
                "rmse": safe_float(values["rmse"]),
                "mspe": safe_float(values["mspe"]),
                "metrics_path": str(metrics_path),
            })

        except Exception as e:
            print(f"读取 {exp_dir.name} 的指标失败: {e}")

    wb = Workbook()
    write_metrics_sheet(wb, rows)
    write_pred_len_avg_sheet(wb, rows)
    wb.save(output_path)
    print(f"\n✅ 指标已导出到: {output_path}")


if __name__ == "__main__":
    main()
